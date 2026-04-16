"""
Build 2025-Topps-Chrome-WWE-Checklist-clean.xlsx from the raw source.
Sheet: Teams_clean
Columns: Card Type | Numbering | Player | Team
"""

import re
import openpyxl
from openpyxl import Workbook

SRC = "/Users/fyrex/Code/CopieStreamlit/wwe check/2025-Topps-Chrome-WWE-Checklist.xlsx"
DST = "/Users/fyrex/Code/CopieStreamlit/wwe check/2025-Topps-Chrome-WWE-Checklist-clean.xlsx"

# ── section-header → Card Type mapping ─────────────────────────────────────
SECTION_MAP = {
    "Base Set":                              "Base",
    "Image Variations Checklist":            "Image Variation",
    "Base Chrome Autographs Gold Refractors Checklist": "Chrome Autograph",
    "Blue Brand Chrome Autographs Checklist":            "Blue Brand Autograph",
    "Future Stars Autographs Checklist":                 "Future Stars Autograph",
    "Hall of Fame Autographs Checklist":                 "Hall of Fame Autograph",
    "Legendary Chrome Autographs Checklist":             "Legendary Autograph",
    "Main Event Autographs Checklist":                   "Main Event Autograph",
    "Marks of Champions Autographs Checklist":           "Marks of Champions Autograph",
    "NXT Chrome Autographs Checklist":                   "NXT Chrome Autograph",
    "Red Brand Chrome Autographs Checklist":             "Red Brand Autograph",
    "1985 Topps Current Checklist":          "1985 Topps Current",
    "1985 Topps Legends Checklist":          "1985 Topps Legends",
    "Allen & Ginter Checklist":              "Allen & Ginter",
    "Embedded Checklist":                    "Embedded",
    "Family Tree Checklist":                 "Family Tree",
    "Feel the Pop Checklist":                "Feel the Pop",
    "Headshots Checklist":                   "Headshots",
    "Helix Checklist":                       "Helix",
    "Hidden Gems Checklist":                 "Hidden Gems",
    "Let's Go! Checklist":                   "Let's Go!",
    "Mania Checklist":                       "Mania",
    "Paradigm Checklist":                    "Paradigm",
    "Persona Checklist":                     "Persona",
    "Shifting Gears Checklist":              "Shifting Gears",
    "Slammy Checklist":                      "Slammy",
    "Tag Team Checklist":                    "Tag Team",
    "The Time Is Now Checklist":             "The Time Is Now",
    "Title Town Checklist":                  "Title Town",
    "Women's Division Checklist":            "Women's Division",
    "WrestleMania Recall Checklist":         "WrestleMania Recall",
}

# Sections whose Team must always be forced (regardless of source column)
FORCE_TEAM = {
    "Hall of Fame Autograph":   "Legend",
    "Legendary Autograph":      "Legend",
    "NXT Chrome Autograph":     "NXT",
    "1985 Topps Legends":       "Legend",
}

# ── helper: clean wrestler name ─────────────────────────────────────────────
def clean_name(raw: str) -> str:
    """
    Strip card codes (e.g. 'CLA-XX ', 'BBA-ANG ') from name strings.
    The source data already has name in column B; just strip extra whitespace
    and trademark symbols, but keep ring-name nicknames in quotes.
    """
    if raw is None:
        return ""
    s = str(raw).strip()
    # Remove trailing trademark symbols
    s = s.replace("™", "").strip()
    return s


# ── helper: clean / normalise team ─────────────────────────────────────────
def clean_team(raw) -> str:
    if raw is None:
        return "WWE"
    s = str(raw).strip()
    # Normalise case variants from source
    mapping = {
        "smackdown": "SmackDown",
        "raw":       "Raw",
        "nxt":       "NXT",
        "legend":    "Legend",
        "wwe":       "WWE",
    }
    return mapping.get(s.lower(), s)


# ── helper: detect card-data rows ──────────────────────────────────────────
# A data row has col0 = card-code (string starting with letters/digits + dash)
# OR col0 = integer (Base Set numbers)
CODE_RE = re.compile(r'^[A-Z0-9].*-.*', re.IGNORECASE)

def is_data_row(row) -> bool:
    c0 = row[0]
    if c0 is None:
        return False
    if isinstance(c0, int):
        return True
    s = str(c0).strip()
    return bool(CODE_RE.match(s))


# ── helper: decide numbering ───────────────────────────────────────────────
def get_numbering(row, card_type: str) -> str:
    """
    Return explicit print-run string (e.g. '/50') only for the
    Chrome Autograph Gold Refractor set, which carries '/50' in col4.
    All other sections: empty string.
    """
    if card_type == "Chrome Autograph":
        val = row[4] if len(row) > 4 else None
        if val and str(val).strip().startswith("/"):
            return str(val).strip()
    return ""


# ── helper: build player / team for multi-person cards ────────────────────
def build_player_team(raw_name: str, raw_team, card_type: str, force_team):
    """
    Some cards (Family Tree, Headshots, Tag Team) list multiple wrestlers.
    - Tag Team: use slash format from source, clean each name
    - Others: keep combined name as-is (after cleaning), pick primary team
    """
    name = clean_name(raw_name)
    team = force_team if force_team else ""

    # If team not forced, derive from source
    if not team:
        # Handle multi-team strings like 'Raw / Smackdown'
        if raw_team and "/" in str(raw_team):
            parts = [clean_team(p.strip()) for p in str(raw_team).split("/")]
            # Use first team, but if it's Legend check second too
            team = parts[0]
        else:
            team = clean_team(raw_team)

    return name, team


# ── main ────────────────────────────────────────────────────────────────────
def main():
    wb_in = openpyxl.load_workbook(SRC)
    ws = wb_in["Full Checklist"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    current_section_key = None   # key in SECTION_MAP
    current_card_type = None

    for i, row in enumerate(rows):
        c0 = row[0]

        # Check if this is a section header
        if c0 is not None and not isinstance(c0, int) and row[1] is None:
            s = str(c0).strip()
            if s in SECTION_MAP:
                current_section_key = s
                current_card_type = SECTION_MAP[s]
            # otherwise it's metadata (card count, odds, parallels) – skip
            continue

        # Data row?
        if current_card_type and is_data_row(row):
            raw_name = row[1]
            raw_team = row[3] if len(row) > 3 else None

            # Skip rows without a name
            if not raw_name:
                continue

            force_team = FORCE_TEAM.get(current_card_type)
            player, team = build_player_team(
                str(raw_name), raw_team, current_card_type, force_team
            )
            numbering = get_numbering(row, current_card_type)

            if player:
                records.append({
                    "Card Type":  current_card_type,
                    "Numbering":  numbering,
                    "Player":     player,
                    "Team":       team,
                })

    # ── write output ────────────────────────────────────────────────────────
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Teams_clean"

    headers = ["Card Type", "Numbering", "Player", "Team"]
    ws_out.append(headers)
    for rec in records:
        numbering = rec["Numbering"] if rec["Numbering"] else ""
        ws_out.append([rec["Card Type"], numbering, rec["Player"], rec["Team"]])

    wb_out.save(DST)

    # ── summary ─────────────────────────────────────────────────────────────
    from collections import Counter
    ct_counts = Counter(r["Card Type"] for r in records)
    print(f"\nTotal rows: {len(records)}")
    print("\nCard Type breakdown:")
    for ct, cnt in sorted(ct_counts.items()):
        print(f"  {ct:<40} {cnt:>4}")
    print(f"\nSaved to: {DST}")


if __name__ == "__main__":
    main()
