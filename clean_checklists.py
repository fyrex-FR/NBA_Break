import os
import re
import shutil

import pandas as pd
from openpyxl import load_workbook

from card_logic import normalize_team_name
from sports_config import DEFAULT_CATEGORY_RULES, DEFAULT_SPORT_KEY, detect_sport_from_filename, get_sport_profile


BASE_BOX_KEYWORDS = ["base", "set", "rookie", "insert", "variation", "parallel"]


def normalize(value):
    return str(value).strip().lower()


def build_box_keywords(sport_profile):
    rules = sport_profile.get("category_rules", DEFAULT_CATEGORY_RULES)
    keywords = set(BASE_BOX_KEYWORDS)
    for values in rules.values():
        for value in values:
            keywords.add(normalize(value))
    return sorted(keywords)


def is_header_row(row):
    joined = " ".join(normalize(v) for v in row if isinstance(v, str))
    return "player" in joined or "team" in joined


def infer_columns(df_raw, team_aliases, box_keywords):
    team_lookup = set(team_aliases.keys())

    team_col = None
    best_ratio = 0
    for col in df_raw.columns:
        col_values = df_raw[col].dropna().tolist()
        if not col_values:
            continue
        matches = sum(1 for v in col_values if normalize(v) in team_lookup)
        ratio = matches / max(len(col_values), 1)
        if ratio > best_ratio:
            best_ratio = ratio
            team_col = col

    player_col = None
    if team_col is not None:
        if team_col - 1 in df_raw.columns:
            player_col = team_col - 1
        elif team_col + 1 in df_raw.columns:
            player_col = team_col + 1

    box_col = None
    best_score = 0
    for col in df_raw.columns:
        if col in (player_col, team_col):
            continue
        col_values = df_raw[col].dropna().tolist()
        if not col_values:
            continue
        score = sum(
            1 for v in col_values
            if isinstance(v, str) and any(k in normalize(v) for k in box_keywords)
        )
        if score > best_score:
            best_score = score
            box_col = col

    if team_col is None:
        team_col = df_raw.columns[-1]
    if player_col is None:
        player_col = df_raw.columns[-2] if len(df_raw.columns) >= 2 else df_raw.columns[0]
    if box_col is None:
        box_col = df_raw.columns[0]

    return player_col, team_col, box_col


def extract_numbering(row):
    for cell in row:
        if isinstance(cell, str):
            match = re.search(r"/\s*(\d+)", cell)
            if match:
                return match.group(1)

    for idx, cell in enumerate(row):
        if isinstance(cell, str) and cell.strip() == "/":
            neighbors = []
            if idx - 1 >= 0:
                neighbors.append(row[idx - 1])
            if idx + 1 < len(row):
                neighbors.append(row[idx + 1])
            for n in neighbors:
                if isinstance(n, (int, float)) and not pd.isna(n):
                    return str(int(n))

    return ""


def process_file(src_path, dst_path, sport_profile):
    df_raw = pd.read_excel(src_path, sheet_name="Teams", header=None, engine="openpyxl")
    df_raw = df_raw.dropna(axis=1, how="all")

    if df_raw.empty:
        return 0

    if is_header_row(df_raw.iloc[0].tolist()):
        df_raw = df_raw.iloc[1:].reset_index(drop=True)

    team_aliases = sport_profile.get("team_aliases", {})
    box_keywords = build_box_keywords(sport_profile)
    player_col, team_col, box_col = infer_columns(df_raw, team_aliases, box_keywords)

    cleaned_rows = []
    for _, row in df_raw.iterrows():
        player = row.get(player_col)
        team = row.get(team_col)
        card_type = row.get(box_col)

        if pd.isna(player) or pd.isna(team):
            continue

        player_str = str(player).strip().rstrip(",")
        team_str = normalize_team_name(team, team_aliases)
        card_str = "" if pd.isna(card_type) else str(card_type).strip()
        numbering = extract_numbering(row.tolist())

        cleaned_rows.append([player_str, team_str, card_str, numbering])

    wb = load_workbook(dst_path)
    if "Teams_clean" in wb.sheetnames:
        del wb["Teams_clean"]
    ws = wb.create_sheet("Teams_clean")
    ws.append(["Player", "Team", "Card Type", "Numbering"])
    for r in cleaned_rows:
        ws.append(r)
    wb.save(dst_path)

    return len(cleaned_rows)


def main():
    src_dir = "/Users/fyrex/antiGravityCode/checklists"
    dst_dir = "/Users/fyrex/antiGravityCode/checklists_clean"

    os.makedirs(dst_dir, exist_ok=True)

    files = [f for f in os.listdir(src_dir) if f.endswith(".xlsx")]
    total_rows = 0

    for fname in sorted(files):
        src_path = os.path.join(src_dir, fname)
        dst_path = os.path.join(dst_dir, fname)
        shutil.copy2(src_path, dst_path)

        sport_key = detect_sport_from_filename(fname, fallback=DEFAULT_SPORT_KEY)
        sport_profile = get_sport_profile(sport_key)

        rows = process_file(src_path, dst_path, sport_profile)
        total_rows += rows
        print(f"{fname}: {rows} lignes ({sport_key})")

    print(f"Fichiers traites: {len(files)}")
    print(f"Total lignes: {total_rows}")


if __name__ == "__main__":
    main()
